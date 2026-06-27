from collections import Counter
from pathlib import Path

from flask import current_app
from openpyxl import load_workbook


def list_diff_reports():
    cache_dir = Path(current_app.config["CACHE_DIR"])
    reports = []
    for path in sorted(cache_dir.glob("*Diff*.xlsx")):
        reports.append({"name": path.stem.replace("_", " "), "path": str(path)})
    return reports


def load_diff_report(path=None, limit=120):
    reports = list_diff_reports()
    if not reports:
        return None
    selected_path = Path(path or reports[0]["path"])
    if not selected_path.exists():
        selected_path = Path(reports[0]["path"])
    workbook = load_workbook(selected_path, read_only=True, data_only=True)
    sheet = workbook["Snapshot diff - 1"] if "Snapshot diff - 1" in workbook.sheetnames else workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    records = []
    counts = Counter()
    for row in rows:
        record = {headers[idx]: value for idx, value in enumerate(row) if idx < len(headers)}
        if not any(record.values()):
            continue
        change_type = str(record.get("Change Type") or "Other").strip() or "Other"
        counts[change_type] += 1
        if len(records) < limit:
            records.append(record)
    return {
        "name": selected_path.stem.replace("_", " "),
        "path": str(selected_path),
        "rows": records,
        "counts": dict(counts),
        "total": sum(counts.values()),
        "reports": reports,
    }
