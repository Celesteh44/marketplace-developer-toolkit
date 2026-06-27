import json
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency at runtime
    load_workbook = None


ATTRIBUTE_ALIASES = {
    "attribute": "name",
    "attribute name": "name",
    "field": "name",
    "field name": "name",
    "name": "name",
    "path": "path",
    "json path": "path",
    "xpath": "path",
    "product type": "product_type",
    "producttype": "product_type",
    "required": "required",
    "is required": "required",
    "data type": "data_type",
    "type": "data_type",
    "enum": "enum_values",
    "enumerations": "enum_values",
    "allowed values": "enum_values",
    "description": "description",
    "definition": "description",
    "example": "example",
    "sample": "example",
}


def parse_spec_file(path):
    path = Path(path)
    if path.suffix.lower() == ".zip":
        return _parse_zip(path)
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _parse_xlsx(path)
    if path.suffix.lower() == ".json":
        return _parse_json(path)
    if path.suffix.lower() == ".xml":
        return _parse_xml(path)
    return []


def _parse_zip(path):
    attributes = []
    with zipfile.ZipFile(path) as archive:
        extract_dir = path.parent / f"{path.stem}_contents"
        extract_dir.mkdir(exist_ok=True)
        for member in archive.namelist():
            suffix = Path(member).suffix.lower()
            if suffix not in {".xlsx", ".xlsm", ".json", ".xml"}:
                continue
            target = extract_dir / Path(member).name
            with archive.open(member) as src, target.open("wb") as dst:
                dst.write(src.read())
            attributes.extend(parse_spec_file(target))
    return attributes


def _parse_xlsx(path):
    if load_workbook is None:
        return []
    workbook = load_workbook(path, read_only=True, data_only=True)
    attributes = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header_idx, mapping = _find_header(rows)
        if header_idx is None:
            continue
        for row in rows[header_idx + 1:]:
            record = _row_to_record(row, mapping, sheet.title)
            if record:
                attributes.append(record)
    return attributes


def _parse_json(path):
    with Path(path).open(encoding="utf-8") as fh:
        data = json.load(fh)
    walmart_product_types = _find_walmart_product_types(data)
    if walmart_product_types:
        attributes = []
        for product_type, schema in walmart_product_types.items():
            _walk_json_schema(schema, [], product_type, attributes, lock_product_type=True)
        return attributes
    attributes = []
    _walk_json_schema(data, [], "General", attributes)
    return attributes


def _parse_xml(path):
    tree = ElementTree.parse(path)
    attributes = []
    _walk_xml(tree.getroot(), [], attributes)
    return attributes


def _find_header(rows):
    for idx, row in enumerate(rows[:20]):
        labels = [_normalize(value) for value in row]
        mapping = {}
        for col_idx, label in enumerate(labels):
            canonical = ATTRIBUTE_ALIASES.get(label)
            if canonical:
                mapping[canonical] = col_idx
        if "name" in mapping or "path" in mapping:
            return idx, mapping
    return None, {}


def _row_to_record(row, mapping, fallback_product_type):
    def value(key):
        idx = mapping.get(key)
        if idx is None or idx >= len(row):
            return ""
        raw = row[idx]
        return "" if raw is None else str(raw).strip()

    name = value("name")
    path = value("path") or name
    if not name and path:
        name = path.split(".")[-1].split("/")[-1]
    if not name:
        return None
    enum_values = _split_enum(value("enum_values"))
    return {
        "product_type": value("product_type") or fallback_product_type,
        "path": path,
        "name": name,
        "required": _is_required(value("required")),
        "data_type": value("data_type"),
        "enum_values": enum_values,
        "description": value("description"),
        "example": value("example"),
        "raw": {},
    }


def _find_walmart_product_types(node):
    if not isinstance(node, dict):
        return {}
    if node.get("title") == "Visible" and isinstance(node.get("properties"), dict):
        properties = node["properties"]
        if len(properties) > 100:
            return properties
    for value in node.values():
        if isinstance(value, dict):
            found = _find_walmart_product_types(value)
            if found:
                return found
        elif isinstance(value, list):
            for item in value:
                found = _find_walmart_product_types(item)
                if found:
                    return found
    return {}


def _walk_json_schema(node, path, product_type, attributes, lock_product_type=False):
    if not isinstance(node, dict):
        return
    title = node.get("title") or node.get("name")
    if title and len(path) <= 1 and not lock_product_type:
        product_type = str(title)
    required = set(node.get("required", [])) if isinstance(node.get("required"), list) else set()
    properties = node.get("properties", {})
    if isinstance(properties, dict):
        for name, child in properties.items():
            child_path = path + [name]
            attributes.append({
                "product_type": product_type,
                "path": ".".join(child_path),
                "name": name,
                "required": name in required,
                "data_type": _json_type(child),
                "enum_values": child.get("enum", []) if isinstance(child, dict) else [],
                "description": child.get("description", "") if isinstance(child, dict) else "",
                "example": _json_example(child),
                "raw": child if isinstance(child, dict) else {},
            })
            _walk_json_schema(child, child_path, product_type, attributes, lock_product_type=lock_product_type)
            if isinstance(child, dict) and isinstance(child.get("items"), dict):
                _walk_json_schema(child["items"], child_path + ["[]"], product_type, attributes, lock_product_type=lock_product_type)


def _walk_xml(element, path, attributes):
    current_path = path + [element.tag]
    children = list(element)
    attributes.append({
        "product_type": path[0] if path else element.tag,
        "path": ".".join(current_path),
        "name": element.tag,
        "required": False,
        "data_type": "object" if children else "string",
        "enum_values": [],
        "description": "",
        "example": (element.text or "").strip()[:120],
        "raw": element.attrib,
    })
    for child in children:
        _walk_xml(child, current_path, attributes)


def _normalize(value):
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _is_required(value):
    return _normalize(value) in {"required", "yes", "y", "true", "1", "mandatory", "required?"}


def _split_enum(value):
    if not value:
        return []
    parts = re.split(r"[,;\n|]", value)
    return [part.strip() for part in parts if part and part.strip()]


def _json_type(node):
    if not isinstance(node, dict):
        return ""
    value = node.get("type", "")
    if isinstance(value, list):
        return ", ".join(value)
    return str(value)


def _stringify(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value)
    return str(value or "")


def _json_example(node):
    if not isinstance(node, dict):
        return ""
    if "example" in node:
        return _stringify(node.get("example"))
    examples = node.get("examples")
    if isinstance(examples, list) and examples:
        return _stringify(examples[0])
    return ""
